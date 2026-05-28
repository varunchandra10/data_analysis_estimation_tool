from __future__ import annotations

import csv as _csv_module
import gzip
import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

from core.config import DATASETS_DIR, DEFAULT_PROJECT
from services.versioning.quality_engine import compute_normalized_quality_score
from utils.hash_utils import generate_sha256
from utils.log_utils import log_calls
from utils.file_utils import safe_json_replace
from utils.dataset_storage import (
    archive_dir,
    cleanup_legacy_root_directories,
    dataset_dir,
    ensure_dataset_layout,
    manifests_dir,
    processed_dir,
    resolve_dataset_name,
    sanitize_dataset_name,
    temp_root,
    versions_dir,
)
from services.dataset_loader import load_dataset, save_dataset



STORAGE_DIR = DATASETS_DIR
DEFAULT_PROJECT_ID = DEFAULT_PROJECT
DATASET_FILENAME = 'dataset.csv'
MANIFEST_SUFFIX = '_manifest.json'


def _project_dir(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    return STORAGE_DIR


def _datasets_root(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    return _project_dir(project_id)


def _versions_dir(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    return _project_dir(project_id)


def _manifests_dir(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    return _project_dir(project_id)


def _logs_dir(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    return _project_dir(project_id)


def _archive_dir(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    return _project_dir(project_id)


def _ai_cache_dir(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    return _project_dir(project_id)


def _sanitize_dataset_name(dataset_name: str) -> str:
    return sanitize_dataset_name(dataset_name)


def _dataset_folder(project_id: str, dataset_name: str) -> Path:
    return dataset_dir(dataset_name)


def _stage_filename(dataset_name: str, stage_name: str, file_extension: str = '.csv') -> str:
    dataset_stem = _sanitize_dataset_name(dataset_name)
    normalized_extension = file_extension if file_extension.startswith('.') else f'.{file_extension}'
    return f'{stage_name}_{dataset_stem}{normalized_extension}'


def _dataset_stage_path(project_id: str, dataset_name: str, stage_name: str, file_extension: str = '.csv') -> Path:
    return processed_dir(dataset_name) / _stage_filename(dataset_name, stage_name, file_extension)


def _ensure_directories(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    project_dir = _project_dir(project_id)
    project_dir.mkdir(parents=True, exist_ok=True)
    cleanup_legacy_root_directories()
    return project_dir


def _timestamp() -> str:
    return datetime.utcnow().isoformat(timespec='seconds') + 'Z'


def _build_stage_manifest(project_id: str, version_name: str, dataset_name: str, dataset_path: Path) -> dict[str, Any]:
    # FIX P2: use fast line-counting instead of loading the full dataframe
    rows, cols = _fast_csv_shape(dataset_path)
    timestamp = _timestamp()

    return {
        'version': version_name,
        'dataset_name': dataset_name,
        'stage_name': version_name,
        'parent': None,
        'children': [],
        'created_at': timestamp,
        'timestamp': timestamp,
        'operations': [],
        'affected_rows': None,
        'lineage': [version_name],
        'lineage_depth': 0,
        'rows': rows,
        'columns': cols,
        'dataset_path': str(dataset_path),
        'size_bytes': int(dataset_path.stat().st_size),
        'hash': generate_sha256(dataset_path),
    }


def _find_raw_dataset_path(project_id: str, dataset_name: str | None = None) -> tuple[Path, str]:
    candidate_datasets = [resolve_dataset_name(dataset_name)] if dataset_name else []
    for folder in _all_dataset_names(project_id):
        if folder not in candidate_datasets:
            candidate_datasets.append(folder)

    for resolved_dataset_name in candidate_datasets:
        dataset_root = processed_dir(resolved_dataset_name)
        for extension in ('.csv', '.xlsx', '.xls', '.csv.gz'):
            raw_path = dataset_root / f'raw_{resolved_dataset_name}{extension}'
            if raw_path.exists():
                return raw_path, resolved_dataset_name

    raise FileNotFoundError('Raw dataset not found for manifest resolution.')


def _resolve_dataset_name_from_source(dataset_source: Any, dataset_name: str | None = None) -> str:
    if dataset_name:
        return resolve_dataset_name(dataset_name)

    if isinstance(dataset_source, pd.DataFrame):
        raise ValueError('dataset_name is required when dataset_source is a DataFrame.')

    return resolve_dataset_name(dataset_source)


def _manifest_path(project_id: str, version_name: str, dataset_name: str | None = None) -> Path:
    target_dataset = dataset_name or _find_dataset_name_for_version(version_name)
    return manifests_dir(target_dataset) / f'{version_name}{MANIFEST_SUFFIX}'


def _version_dataset_path(project_id: str, version_name: str, dataset_name: str | None = None) -> Path:
    target_dataset = dataset_name or _find_dataset_name_for_version(version_name)
    return versions_dir(target_dataset) / version_name / DATASET_FILENAME


def _all_dataset_names(project_id: str = DEFAULT_PROJECT_ID) -> list[str]:
    _ensure_directories(project_id)
    names: list[str] = []
    for folder in sorted(_datasets_root(project_id).glob('*')):
        if folder.is_dir():
            ensure_dataset_layout(folder.name)
            names.append(folder.name)
    return names


def _find_manifest_path(version_name: str, project_id: str = DEFAULT_PROJECT_ID) -> Path:
    for dataset_name in _all_dataset_names(project_id):
        manifest_path = manifests_dir(dataset_name) / f'{version_name}{MANIFEST_SUFFIX}'
        if manifest_path.exists():
            return manifest_path
    raise FileNotFoundError(f'Manifest not found for version: {version_name}')


def _find_dataset_name_for_version(version_name: str, project_id: str = DEFAULT_PROJECT_ID) -> str:
    manifest_path = _find_manifest_path(version_name, project_id)
    return manifest_path.parent.parent.parent.name


def _normalize_dataset_source(dataset_source: Any, target_path: Path) -> Path:
    target_path.parent.mkdir(parents=True, exist_ok=True)

    if isinstance(dataset_source, pd.DataFrame):
        save_dataset(dataset_source, target_path)
        return target_path

    source_path = Path(dataset_source)
    if not source_path.exists():
        raise FileNotFoundError(f'Dataset source not found: {source_path}')

    if source_path.is_dir():
        raise IsADirectoryError(f'Dataset source must be a file: {source_path}')

    target_suffix = target_path.suffix.lower()
    source_suffix = source_path.suffix.lower()

    if target_suffix in {'.xlsx', '.xls'}:
        if source_suffix in {'.xlsx', '.xls'}:
            shutil.copy2(source_path, target_path)
            return target_path
        if source_suffix in {'.csv', '.gz'} or source_path.name.endswith('.csv.gz'):
            df = load_dataset(source_path, optimize=False)
            save_dataset(df, target_path)
            return target_path

    if target_suffix in {'.csv', '.gz'}:
        if source_suffix in {'.csv', '.gz'} or source_path.name.endswith('.csv.gz'):
            shutil.copy2(source_path, target_path)
            return target_path
        if source_suffix in {'.xlsx', '.xls'}:
            df = load_dataset(source_path, optimize=False)
            save_dataset(df, target_path)
            return target_path

    shutil.copy2(source_path, target_path)
    return target_path



def _is_valid_dataset_file(path: Path) -> bool:
    """Lightweight extension check — no file read. Use instead of load_dataset for existence validation."""
    name = path.name.lower()
    suffix = path.suffix.lower()
    return suffix in {'.csv', '.xlsx', '.xls'} or name.endswith('.csv.gz')


def _fast_csv_shape(path: Path) -> tuple[int, int]:
    """Count rows and columns without loading into pandas. Returns (row_count, col_count).
    Supports .csv, .csv.gz, .xlsx, .xls. Falls back to (0, 0) on any error."""
    name = path.name.lower()
    suffix = path.suffix.lower()
    try:
        if name.endswith('.csv.gz'):
            with gzip.open(path, 'rt', encoding='utf-8', errors='replace', newline='') as f:
                reader = _csv_module.reader(f)
                header = next(reader, [])
                rows = sum(1 for _ in reader)
            return rows, len(header)
        if suffix == '.csv':
            with open(path, 'r', encoding='utf-8', errors='replace', newline='') as f:
                reader = _csv_module.reader(f)
                header = next(reader, [])
                rows = sum(1 for _ in reader)
            return rows, len(header)
        if suffix in {'.xlsx', '.xls'}:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                cols = ws.max_column or 0
                rows = max((ws.max_row or 1) - 1, 0)
                wb.close()
                return rows, cols
            except Exception:
                # openpyxl unavailable — fall through to pandas header-only read
                df_head = pd.read_excel(path, nrows=0)
                cols = len(df_head.columns)
                # row count still needs a full read for Excel
                rows = len(pd.read_excel(path))
                return rows, cols
    except Exception:
        pass
    return 0, 0


def _read_dataset(dataset_path: Path) -> pd.DataFrame:
    # Delegate to centralized dataset loader which handles formats and optimization
    return load_dataset(dataset_path, optimize=True)


def read_dataset_file(dataset_path: str | Path) -> pd.DataFrame:
    return _read_dataset(Path(dataset_path))


def preview_dataset_file(dataset_path: str | Path, rows: int = 5) -> dict[str, Any]:
    df = read_dataset_file(dataset_path)
    return {
        'rows': int(df.shape[0]),
        'columns': int(df.shape[1]),
        'preview': safe_json_replace(df.head(rows)),
        'columns_list': safe_json_replace(list(df.columns)),
    }


def save_stage_dataset(
    dataset_source: Any,
    dataset_name: str,
    stage_name: str,
    file_extension: str | None = None,
    project_id: str = DEFAULT_PROJECT_ID,
) -> Path:
    _ensure_directories(project_id)

    resolved_dataset_name = resolve_dataset_name(dataset_name)
    ensure_dataset_layout(resolved_dataset_name)

    if file_extension is None:
        if isinstance(dataset_source, pd.DataFrame):
            file_extension = '.csv'
        else:
            source_path = Path(dataset_source)
            if source_path.name.endswith('.csv.gz'):
                file_extension = '.csv'
            elif source_path.suffix.lower() in {'.csv', '.xlsx', '.xls'}:
                file_extension = source_path.suffix.lower()
            else:
                file_extension = '.csv'

    target_path = _dataset_stage_path(project_id, resolved_dataset_name, stage_name, file_extension)
    return _normalize_dataset_source(dataset_source, target_path)


def _stage_order(stage_name: str) -> int:
    order = {
        'raw': 0,
        'preprocessing': 1,
        'clean': 1,
        'cleaned': 1,
        'outlier': 2,
        'outlier_detected': 2,
        'outliers': 2,
        'validation': 3,
        'validated': 3,
        'weighting': 4,
        'estimation': 4,
        'estimated': 4,
        'deduped': 5,
        'deduplicated': 5,
        'ai': 6,
        'report_generation': 7,
    }
    return order.get(stage_name, 99)


def get_dataset_folders(project_id: str = DEFAULT_PROJECT_ID) -> list[dict[str, Any]]:
    _ensure_directories(project_id)
    folders: list[dict[str, Any]] = []

    for folder in sorted(_datasets_root(project_id).glob('*')):
        if not folder.is_dir():
            continue
        ensure_dataset_layout(folder.name)
        current_processed_dir = processed_dir(folder.name)
        files: set[str] = {path.name for path in sorted(current_processed_dir.iterdir()) if path.is_file()}
        for version_dir in sorted(versions_dir(folder.name).iterdir()):
            if version_dir.is_dir() and (version_dir / DATASET_FILENAME).exists():
                files.add(f'{version_dir.name}.csv')

        sorted_files = sorted(files, key=lambda name: (_stage_order(name.split('_', 1)[0]), name))
        folders.append({
            'dataset_name': folder.name,
            'folder_path': str(folder),
            'files': sorted_files,
        })

    return folders


def get_dataset_files(
    dataset_name: str,
    project_id: str = DEFAULT_PROJECT_ID,
) -> list[dict[str, Any]]:
    folder = _dataset_folder(project_id, dataset_name)
    if not folder.exists():
        return []

    ensure_dataset_layout(folder.name)
    current_processed_dir = processed_dir(folder.name)
    files: list[dict[str, Any]] = []
    for path in sorted(current_processed_dir.iterdir()):
        if not path.is_file():
            continue
        files.append({
            'file_name': path.name,
            'file_path': str(path),
            'stage': path.name.split('_', 1)[0],
            'size_bytes': path.stat().st_size,
        })

    for version_dir in sorted(versions_dir(folder.name).iterdir()):
        if not version_dir.is_dir():
            continue

        dataset_path = version_dir / DATASET_FILENAME
        if not dataset_path.exists():
            continue

        manifest = _load_manifest_optional(project_id, version_dir.name, folder.name)
        stage_name = manifest.get('stage_name') or version_dir.name.split('_', 1)[-1]
        files.append({
            'file_name': f'{version_dir.name}.csv',
            'file_path': str(dataset_path),
            'stage': stage_name,
            'version': version_dir.name,
            'size_bytes': dataset_path.stat().st_size,
        })

    files.sort(key=lambda item: (_stage_order(item.get('stage', '')), item.get('version', ''), item['file_name']))
    return files


@log_calls
def delete_dataset_file(dataset_name: str, file_name: str, project_id: str = DEFAULT_PROJECT_ID) -> Path:
    folder = _dataset_folder(project_id, dataset_name)
    if not folder.exists():
        raise FileNotFoundError(f'Dataset folder not found: {folder}')

    ensure_dataset_layout(folder.name)
    target = processed_dir(folder.name) / file_name
    if not target.exists():
        raise FileNotFoundError(f'Dataset file not found: {target}')

    target.unlink()
    return target


def get_versioning_dashboard(project_id: str = DEFAULT_PROJECT_ID) -> dict[str, Any]:
    return {
        'project_id': project_id,
        'datasets': get_dataset_folders(project_id),
    }


def _extract_version_rank(version_name: str) -> tuple[int, str]:
    prefix = version_name.split('_', 1)[0]
    if prefix.startswith('v') and prefix[1:].isdigit():
        return int(prefix[1:]), version_name
    if version_name.startswith('v') and version_name[1:].isdigit():
        return int(version_name[1:]), version_name
    return -1, version_name


def _load_manifest(project_id: str, version_name: str, dataset_name: str | None = None) -> dict[str, Any]:
    if version_name == 'raw':
        raw_path, resolved_dataset_name = _find_raw_dataset_path(project_id, dataset_name)
        return _build_stage_manifest(project_id, version_name, resolved_dataset_name, raw_path)

    manifest_path = _manifest_path(project_id, version_name, dataset_name)
    if not manifest_path.exists():
        raise FileNotFoundError(f'Manifest not found for version: {version_name}')
    with open(manifest_path, 'r', encoding='utf-8') as handle:
        return json.load(handle)


def _load_manifest_optional(project_id: str, version_name: str, dataset_name: str | None = None) -> dict[str, Any]:
    try:
        if dataset_name:
            manifest_path = _manifest_path(project_id, version_name, dataset_name)
            if manifest_path.exists():
                with open(manifest_path, 'r', encoding='utf-8') as handle:
                    return json.load(handle)
        return _load_manifest(project_id, version_name, dataset_name)
    except FileNotFoundError:
        return {}


def _resolve_version_dataset_path(project_id: str, version_name: str, dataset_name: str | None = None) -> tuple[Path, str | None]:
    """Resolve the filesystem path for a version dataset.

    P2 FIX: replaced load_dataset() existence probes with Path.exists() + extension check.
    No dataframe is loaded; only filesystem metadata is inspected.
    """
    candidate_datasets: list[str] = []
    if dataset_name:
        candidate_datasets.append(resolve_dataset_name(dataset_name))

    for folder in _all_dataset_names(project_id):
        if folder not in candidate_datasets:
            candidate_datasets.append(folder)

    for resolved_dataset_name in candidate_datasets:
        # Check canonical versions/<version>/dataset.csv path
        version_path = _version_dataset_path(project_id, version_name, resolved_dataset_name)
        if version_path.exists() and _is_valid_dataset_file(version_path):
            return version_path, resolved_dataset_name

        # Check processed stage files with known extensions
        for extension in ('.csv', '.xlsx', '.xls', '.csv.gz'):
            stage_path = processed_dir(resolved_dataset_name) / f'{version_name}_{resolved_dataset_name}{extension}'
            if stage_path.exists() and _is_valid_dataset_file(stage_path):
                return stage_path, resolved_dataset_name

        # Glob fallback: find any file matching version_name prefix
        for path in sorted(processed_dir(resolved_dataset_name).glob(f'{version_name}_*')):
            if path.is_file() and _is_valid_dataset_file(path):
                return path, resolved_dataset_name

    raise FileNotFoundError(f'Version dataset not found: {version_name}')


def _load_version_frame(project_id: str, version_name: str, dataset_name: str | None = None) -> tuple[pd.DataFrame, dict[str, Any], str]:
    resolved_dataset_name = resolve_dataset_name(dataset_name) if dataset_name else None
    dataset_path, resolved_dataset_name = _resolve_version_dataset_path(project_id, version_name, resolved_dataset_name)
    manifest = _load_manifest_optional(project_id, version_name, resolved_dataset_name)
    resolved_dataset_name = resolved_dataset_name or manifest.get('dataset_name') or dataset_path.parent.parent.parent.name
    return read_dataset_file(dataset_path), manifest, resolved_dataset_name


def _count_missing_values(df: pd.DataFrame) -> int:
    return int(df.isna().sum().sum())


def _count_duplicate_rows(df: pd.DataFrame) -> int:
    return int(df.duplicated().sum())


def _count_outlier_rows(df: pd.DataFrame) -> int:
    numeric_df = df.select_dtypes(include='number')
    if numeric_df.empty:
        return 0

    outlier_mask = pd.Series(False, index=df.index)
    for column in numeric_df.columns:
        series = numeric_df[column].dropna()
        if series.empty:
            continue
        q1 = series.quantile(0.25)
        q3 = series.quantile(0.75)
        iqr = q3 - q1
        if pd.isna(iqr) or iqr == 0:
            continue
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        outlier_mask = outlier_mask | (numeric_df[column] < lower) | (numeric_df[column] > upper)

    return int(outlier_mask.sum())


def _validation_violation_count(manifest: dict[str, Any], df: pd.DataFrame) -> int:
    validation = manifest.get('validation')
    if isinstance(validation, dict):
        return int(validation.get('total_violations', validation.get('failed_rules', 0)) or 0)

    null_count = _count_missing_values(df)
    if null_count > 0:
        return null_count

    return 0


def _improvement(before: int, after: int) -> dict[str, Any]:
    delta = before - after
    percent = round((delta / before) * 100, 2) if before else (100.0 if after == 0 else 0.0)
    return {
        'before': before,
        'after': after,
        'delta': delta,
        'reduction_percent': percent,
        'improved': delta > 0,
    }


@log_calls
def compare_versions(
    left_version: str,
    right_version: str,
    project_id: str = DEFAULT_PROJECT_ID,
    dataset_name: str | None = None,
) -> dict[str, Any]:
    left_df, left_manifest, resolved_dataset_name = _load_version_frame(project_id, left_version, dataset_name)
    right_df, right_manifest, _ = _load_version_frame(project_id, right_version, resolved_dataset_name)

    left_metrics = {
        'rows': int(left_df.shape[0]),
        'columns': int(left_df.shape[1]),
        'missing_values': _count_missing_values(left_df),
        'duplicate_rows': _count_duplicate_rows(left_df),
        'outlier_rows': _count_outlier_rows(left_df),
        'validation_violations': _validation_violation_count(left_manifest, left_df),
    }
    right_metrics = {
        'rows': int(right_df.shape[0]),
        'columns': int(right_df.shape[1]),
        'missing_values': _count_missing_values(right_df),
        'duplicate_rows': _count_duplicate_rows(right_df),
        'outlier_rows': _count_outlier_rows(right_df),
        'validation_violations': _validation_violation_count(right_manifest, right_df),
    }

    comparison = {
        'missing_reduction': _improvement(left_metrics['missing_values'], right_metrics['missing_values']),
        'duplicate_reduction': _improvement(left_metrics['duplicate_rows'], right_metrics['duplicate_rows']),
        'outlier_reduction': _improvement(left_metrics['outlier_rows'], right_metrics['outlier_rows']),
        'validation_improvement': _improvement(left_metrics['validation_violations'], right_metrics['validation_violations']),
    }

    return {
        'status': 'success',
        'dataset_name': resolved_dataset_name,
        'left_version': {
            'version': left_version,
            'manifest_path': str(_manifest_path(project_id, left_version, resolved_dataset_name)),
            'dataset_path': str(_version_dataset_path(project_id, left_version, resolved_dataset_name)),
            'metadata': {
                'timestamp': left_manifest.get('timestamp', left_manifest.get('created_at')),
                'stage_name': left_manifest.get('stage_name'),
                'parent': left_manifest.get('parent'),
            },
            'metrics': left_metrics,
        },
        'right_version': {
            'version': right_version,
            'manifest_path': str(_manifest_path(project_id, right_version, resolved_dataset_name)),
            'dataset_path': str(_version_dataset_path(project_id, right_version, resolved_dataset_name)),
            'metadata': {
                'timestamp': right_manifest.get('timestamp', right_manifest.get('created_at')),
                'stage_name': right_manifest.get('stage_name'),
                'parent': right_manifest.get('parent'),
            },
            'metrics': right_metrics,
        },
        'comparison': comparison,
        'summary': {
            'compare': f'{left_version} vs {right_version}',
            'dataset_name': resolved_dataset_name,
            'overall_improvement': round(
                sum(item['delta'] for item in comparison.values()),
                2,
            ),
        },
    }


def _build_lineage(project_id: str, dataset_name: str, version_name: str, parent: str | None) -> list[str]:
    lineage: list[str] = [version_name]
    current_parent = parent

    while current_parent:
        lineage.insert(0, current_parent)
        try:
            parent_manifest = _load_manifest(project_id, current_parent, dataset_name)
        except FileNotFoundError:
            break
        current_parent = parent_manifest.get('parent')

    return lineage


def _append_child_version(project_id: str, dataset_name: str, parent_version: str, child_version: str) -> None:
    try:
        parent_manifest_path = _manifest_path(project_id, parent_version, dataset_name)
    except FileNotFoundError:
        return

    if not parent_manifest_path.exists():
        return

    with open(parent_manifest_path, 'r', encoding='utf-8') as handle:
        parent_manifest = json.load(handle)

    children = list(parent_manifest.get('children', []))
    if child_version not in children:
        children.append(child_version)

    parent_manifest['children'] = children
    parent_manifest['latest_child'] = child_version

    with open(parent_manifest_path, 'w', encoding='utf-8') as handle:
        json.dump(parent_manifest, handle, indent=4)


@log_calls
def create_project(project_id: str = DEFAULT_PROJECT_ID) -> Path:
    return _ensure_directories(project_id)


@log_calls
def compress_dataset_folder(
    dataset_name: str,
    project_id: str = DEFAULT_PROJECT_ID,
    output_path: str | Path | None = None,
    remove_source: bool = False,
) -> Path:
    folder = _dataset_folder(project_id, dataset_name)
    if not folder.exists():
        raise FileNotFoundError(f'Dataset folder not found: {folder}')

    ensure_dataset_layout(folder.name)
    current_archive_dir = archive_dir(folder.name)

    target_path = Path(output_path) if output_path else current_archive_dir / f'{folder.name}.zip'
    target_path.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(target_path, 'w', compression=zipfile.ZIP_DEFLATED) as zip_handle:
        for path in folder.rglob('*'):
            if path.is_file() and path != target_path:
                zip_handle.write(path, arcname=path.relative_to(folder))

    if remove_source:
        shutil.rmtree(folder)

    return target_path


@log_calls
def save_raw_dataset(
    dataset_source: Any,
    project_id: str = DEFAULT_PROJECT_ID,
    filename: str = DATASET_FILENAME,
) -> Path:
    return save_stage_dataset(
        dataset_source=dataset_source,
        dataset_name=filename,
        stage_name='raw',
        project_id=project_id,
    )


@log_calls
def save_manifest(
    version_name: str,
    project_id: str = DEFAULT_PROJECT_ID,
    dataset_name: str | None = None,
    parent: str | None = None,
    stage_name: str | None = None,
    operations: Iterable[str] | None = None,
    affected_rows: int | None = None,
    dataset_path: str | Path | None = None,
    extra: dict[str, Any] | None = None,
) -> Path:
    _ensure_directories(project_id)
    resolved_dataset_name = _resolve_dataset_name_from_source(dataset_path or version_name, dataset_name)
    ensure_dataset_layout(resolved_dataset_name)

    lineage = _build_lineage(project_id, resolved_dataset_name, version_name, parent)
    timestamp = _timestamp()

    manifest: dict[str, Any] = {
        'version': version_name,
        'dataset_name': resolved_dataset_name,
        'stage_name': stage_name or version_name,
        'parent': parent,
        'children': [],
        'created_at': timestamp,
        'timestamp': timestamp,
        'operations': list(operations or []),
        'affected_rows': affected_rows,
        'lineage': lineage,
        'lineage_depth': max(len(lineage) - 1, 0),
    }

    if dataset_path is not None:
        dataset_file = Path(dataset_path)
        if dataset_file.exists():
            # P2 FIX: use fast line-counting instead of loading the full dataframe.
            # This saves a full pd.read_csv on every version save (6+ times per pipeline run).
            rows, cols = _fast_csv_shape(dataset_file)
            manifest.update({
                'rows': rows,
                'columns': cols,
                'dataset_path': str(dataset_file),
                'size_bytes': int(dataset_file.stat().st_size),
                'hash': generate_sha256(dataset_file),
            })

    if extra:
        manifest.update(extra)

    manifest_path = _manifest_path(project_id, version_name, resolved_dataset_name)
    with open(manifest_path, 'w', encoding='utf-8') as handle:
        json.dump(manifest, handle, indent=4)

    if parent:
        _append_child_version(project_id, resolved_dataset_name, parent, version_name)

    return manifest_path


@log_calls
def create_version(
    dataset_source: Any,
    version_name: str,
    project_id: str = DEFAULT_PROJECT_ID,
    dataset_name: str | None = None,
    parent: str | None = None,
    stage_name: str | None = None,
    operations: Iterable[str] | None = None,
    affected_rows: int | None = None,
    extra_manifest: dict[str, Any] | None = None,
) -> dict[str, Any]:
    _ensure_directories(project_id)
    resolved_dataset_name = _resolve_dataset_name_from_source(dataset_source, dataset_name)
    ensure_dataset_layout(resolved_dataset_name)

    version_dir = versions_dir(resolved_dataset_name) / version_name
    dataset_path = version_dir / DATASET_FILENAME

    if version_dir.exists() or _manifest_path(project_id, version_name, resolved_dataset_name).exists():
        raise FileExistsError(f'Version already exists: {version_name}')

    _normalize_dataset_source(dataset_source, dataset_path)

    manifest_path = None
    try:
        manifest_path = save_manifest(
            version_name=version_name,
            project_id=project_id,
            dataset_name=resolved_dataset_name,
            parent=parent,
            stage_name=stage_name,
            operations=operations,
            affected_rows=affected_rows,
            dataset_path=dataset_path,
            extra=extra_manifest,
        )

        from utils.file_utils import resolve_safe_path
        safe_dataset_path = resolve_safe_path(dataset_path)
        checksum = generate_sha256(safe_dataset_path)

        from core.database import SessionLocal
        from models.dataset_model import Dataset
        from models.version_model import Version

        db = SessionLocal()
        try:
            ds = db.query(Dataset).filter(Dataset.dataset_name == resolved_dataset_name).first()
            if not ds:
                from models.project_model import Project
                proj = db.query(Project).filter(Project.id == 1).first()
                project_id = proj.id if proj else None

                ds = Dataset(
                    project_id=project_id,
                    dataset_name=resolved_dataset_name,
                    file_path=str(safe_dataset_path),
                    version=version_name,
                    status="active"
                )
                db.add(ds)
                db.commit()
                db.refresh(ds)
            else:
                ds.file_path = str(safe_dataset_path)
                ds.version = version_name
                ds.status = stage_name or version_name
                ds.checksum = checksum
                try:
                    ds.size = safe_dataset_path.stat().st_size
                except OSError:
                    pass
                db.commit()
                db.refresh(ds)


            db_version = Version(
                dataset_id=ds.id,
                version_name=version_name,
                stage=stage_name or version_name,
                file_path=str(safe_dataset_path),
                checksum=checksum,
                parent_version=parent,
                lineage_metadata=json.dumps(_build_lineage(project_id, resolved_dataset_name, version_name, parent))
            )
            db.add(db_version)
            db.commit()
            db.refresh(db_version)
        finally:
            db.close()

    except Exception as e:
        if version_dir.exists():
            shutil.rmtree(version_dir)
        if manifest_path and manifest_path.exists():
            try:
                manifest_path.unlink()
            except Exception:
                pass
        raise e

    return {
        'project_id': project_id,
        'dataset_name': resolved_dataset_name,
        'version': version_name,
        'parent': parent,
        'stage_name': stage_name or version_name,
        'affected_rows': affected_rows,
        'lineage': _build_lineage(project_id, resolved_dataset_name, version_name, parent),
        'version_dir': str(version_dir),
        'dataset_path': str(dataset_path),
        'manifest_path': str(manifest_path),
    }



@log_calls
def list_versions(project_id: str = DEFAULT_PROJECT_ID) -> list[dict[str, Any]]:
    _ensure_directories(project_id)

    versions: list[dict[str, Any]] = []
    for dataset_name in _all_dataset_names(project_id):
        for manifest_path in sorted(manifests_dir(dataset_name).glob(f'*{MANIFEST_SUFFIX}')):
            with open(manifest_path, 'r', encoding='utf-8') as handle:
                manifest = json.load(handle)
            version_name = manifest.get('version') or manifest_path.stem.replace('_manifest', '')
            versions.append({
                'version': version_name,
                'dataset_name': dataset_name,
                'manifest_path': str(manifest_path),
                'dataset_path': str(_version_dataset_path(project_id, version_name, dataset_name)),
                'created_at': manifest.get('created_at'),
                'timestamp': manifest.get('timestamp', manifest.get('created_at')),
                'parent': manifest.get('parent'),
                'children': manifest.get('children', []),
                'lineage': manifest.get('lineage', [version_name]),
                'operations': manifest.get('operations', []),
                'affected_rows': manifest.get('affected_rows'),
                'hash': manifest.get('hash'),
                'stage_name': manifest.get('stage_name'),
            })

    versions.sort(key=lambda item: (
        _extract_version_rank(item['version'])[0],
        item.get('created_at') or '',
        item['version'],
    ))
    return versions


@log_calls
def get_latest_version(project_id: str = DEFAULT_PROJECT_ID) -> dict[str, Any] | None:
    versions = list_versions(project_id)
    if not versions:
        return None
    return versions[-1]


@log_calls
def rollback_version(
    version_name: str,
    project_id: str = DEFAULT_PROJECT_ID,
    dataset_name: str | None = None,
    restore_to: str | Path | None = None,
) -> Path:
    """Restore the active dataset to a specific historical version snapshot.

    Steps:
      1. Locate version in DB + filesystem
      2. Validate ownership exists
      3. Copy version file → active dataset path (or restore_to)
      4. Update Dataset DB row (version, status, checksum, size)
      5. DB commit (atomic — raises on failure, rolls back)
      6. Refresh manifest with rollback_event annotation
      7. Invalidate stale AI cache for the dataset
      8. Write audit log entry
    Steps 6–8 are best-effort: a failure does not undo the completed rollback.
    """
    from utils.log_utils import logger

    _ensure_directories(project_id)

    # ── Steps 1–5: core rollback (atomic) ────────────────────────────────────
    from models.version_model import Version
    from core.database import SessionLocal
    db = SessionLocal()
    resolved_dataset_name: str | None = None
    previous_version: str | None = None
    try:
        db_version = db.query(Version).filter(Version.version_name == version_name).first()
        if not db_version:
            raise FileNotFoundError(f'Version dataset not found in DB: {version_name}')
        dataset_path = Path(db_version.file_path)

        if not dataset_path.exists():
            raise FileNotFoundError(f'Version dataset file not found on disk: {version_name} → {dataset_path}')

        if restore_to is None:
            from models.dataset_model import Dataset
            ds = db.query(Dataset).filter(Dataset.id == db_version.dataset_id).first()
            if not ds:
                raise FileNotFoundError(f'Associated Dataset not found for ID: {db_version.dataset_id}')

            # Capture pre-rollback state for audit log
            previous_version = ds.version
            resolved_dataset_name = ds.dataset_name

            active_file_path = Path(ds.file_path)
            active_file_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(dataset_path, active_file_path)

            ds.version = version_name
            ds.status = db_version.stage or 'rolled_back'
            ds.checksum = db_version.checksum
            ds.size = dataset_path.stat().st_size
            db.commit()

            restore_path = active_file_path
        else:
            restore_path = Path(restore_to)
            restore_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(dataset_path, restore_path)
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()

    # ── Steps 6–8: best-effort post-rollback housekeeping ────────────────────
    # Failure here does NOT undo the completed rollback — we log and continue.

    # Step 6: Refresh manifest with rollback_event annotation
    if resolved_dataset_name and restore_to is None:
        try:
            rollback_timestamp = _timestamp()
            save_manifest(
                version_name=version_name,
                project_id=project_id,
                dataset_name=resolved_dataset_name,
                dataset_path=restore_path,
                extra={
                    'rollback_event': {
                        'rolled_back_at': rollback_timestamp,
                        'rolled_back_from': previous_version,
                        'restore_target': str(restore_path),
                    }
                },
            )
        except Exception as manifest_err:
            logger.warning(
                f'rollback_version: manifest refresh failed for {version_name} '
                f'(non-critical, rollback completed): {manifest_err}',
                exc_info=True,
            )

    # Step 7: Invalidate stale AI cache
    if resolved_dataset_name:
        try:
            from services.ai.cache_engine import invalidate_ai_cache
            invalidate_ai_cache(resolved_dataset_name)
        except Exception as cache_err:
            logger.warning(
                f'rollback_version: AI cache invalidation failed for {resolved_dataset_name} '
                f'(non-critical, rollback completed): {cache_err}',
                exc_info=True,
            )

    # Step 8: Audit log
    try:
        from services.audit_log_service import record_audit_log
        record_audit_log(
            dataset_name=resolved_dataset_name,
            action='rollback',
            details={
                'version_name': version_name,
                'rolled_back_from': previous_version,
                'restore_path': str(restore_path),
                'description': f'Dataset rolled back to version {version_name}.',
            },
            status='success',
        )
    except Exception as audit_err:
        logger.warning(
            f'rollback_version: audit log failed for {version_name} '
            f'(non-critical, rollback completed): {audit_err}',
            exc_info=True,
        )

    return restore_path




@log_calls
def get_manifest(project_id: str, version_name: str, dataset_name: str | None = None) -> dict[str, Any]:
    return _load_manifest(project_id, version_name, dataset_name)


@log_calls
def get_version_path(project_id: str, version_name: str) -> Path:
    return _version_dataset_path(project_id, version_name)


@log_calls
def compute_quality_score(
    version_name: str,
    project_id: str = DEFAULT_PROJECT_ID,
    dataset_name: str | None = None,
) -> dict[str, Any]:
    """Compute deterministic normalized quality score for a specific version snapshot."""
    df, manifest, resolved_dataset_name = _load_version_frame(project_id, version_name, dataset_name)

    rows = int(df.shape[0])
    cols = int(df.shape[1])
    total_cells = rows * cols if rows and cols else 1

    missing = _count_missing_values(df)
    duplicates = _count_duplicate_rows(df)
    outliers = _count_outlier_rows(df)
    validation = _validation_violation_count(manifest, df)

    # percent metrics (0..100)
    missing_pct = round((missing / total_cells) * 100, 2) if total_cells else 0.0
    duplicate_pct = round((duplicates / rows) * 100, 2) if rows else 0.0
    outlier_pct = round((outliers / rows) * 100, 2) if rows else 0.0
    validation_pct = round((validation / rows) * 100, 2) if rows else 0.0

    normalized_quality = compute_normalized_quality_score(
        missing_percent=missing_pct,
        duplicate_percent=duplicate_pct,
        outlier_percent=outlier_pct,
        validation_violation_percent=validation_pct,
    )

    return {
        'status': 'success',
        'version': version_name,
        'dataset_name': resolved_dataset_name,
        'rows': rows,
        'columns': cols,
        'metrics': {
            'missing': missing,
            'missing_pct': missing_pct,
            'duplicates': duplicates,
            'duplicate_pct': duplicate_pct,
            'outliers': outliers,
            'outlier_pct': outlier_pct,
            'validation_violations': validation,
            'validation_pct': validation_pct,
        },
        'score': normalized_quality['score'],
        'grade': normalized_quality['grade'],
        'breakdown': {
            'weights': normalized_quality['weights'],
            'components': normalized_quality['components'],
            'inputs': normalized_quality['inputs'],
        },
    }
